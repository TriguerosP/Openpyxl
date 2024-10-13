import openpyxl

def crear_archivo_excel(nombre_archivo):
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = "Gastos"

    hoja.append(["Fecha", "Descripción", "Monto"])
    workbook.save(nombre_archivo)

def agregar_gasto(nombre_archivo, fecha, descripcion, monto):
    workbook = openpyxl.load_workbook(nombre_archivo)
    hoja = workbook["Gastos"]
    hoja.append([fecha, descripcion, monto])
    workbook.save(nombre_archivo)

def generar_informe(nombre_archivo):
    workbook = openpyxl.load_workbook(nombre_archivo)
    hoja = workbook["Gastos"]

    gastos = []
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        gastos.append(fila)

    if not gastos:
        print("No hay gastos registrados.")
        return

    total_gastos = sum(gasto[2] for gasto in gastos)
    num_gastos = len(gastos)
    gasto_mas_caro = max(gastos, key=lambda x: x[2])
    gasto_mas_barato = min(gastos, key=lambda x: x[2])

    print(f"Número total de gastos: {num_gastos}")
    print(f"Gasto más caro: {gasto_mas_caro[0]} - {gasto_mas_caro[1]} - ${gasto_mas_caro[2]:.2f}")
    print(f"Gasto más barato: {gasto_mas_barato[0]} - {gasto_mas_barato[1]} - ${gasto_mas_barato[2]:.2f}")
    print(f"Monto total de gastos: ${total_gastos:.2f}")

    hoja_informe = workbook.create_sheet(title="Resumen")
    hoja_informe.append(["Número total de gastos", num_gastos])
    hoja_informe.append(["Gasto más caro", gasto_mas_caro[0], gasto_mas_caro[1], gasto_mas_caro[2]])
    hoja_informe.append(["Gasto más barato", gasto_mas_barato[0], gasto_mas_barato[1], gasto_mas_barato[2]])
    hoja_informe.append(["Monto total de gastos", total_gastos])
    workbook.save(nombre_archivo)
    print("Informe de gastos guardado en 'informe_gastos.xlsx'.")

def main():
    nombre_archivo = "informe_gastos.xlsx"
    try:
        crear_archivo_excel(nombre_archivo)
    except FileExistsError:
        pass

    while True:
        fecha = input("Ingrese la fecha del gasto (YYYY-MM-DD): ")
        descripcion = input("Ingrese la descripción del gasto: ")
        try:
            monto = float(input("Ingrese el monto del gasto: "))
        except ValueError:
            print("Por favor, ingrese un monto válido.")
            continue

        agregar_gasto(nombre_archivo, fecha, descripcion, monto)

        continuar = input("¿Desea agregar otro gasto? (s/n): ").lower()
        if continuar != 's':
            break

    generar_informe(nombre_archivo)

if __name__ == "__main__":
    main()
